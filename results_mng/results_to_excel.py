import pandas as pd
import numpy as np 
import json
from openpyxl import load_workbook

def devolver_text(text):
    return str(text)

def write_results(json_file,excel_file):
    
    with open(json_file) as input_file:
        json_info=json.load(input_file)
    df = pd.read_excel(excel_file, converters={"Case identifier": devolver_text})
    workbook = load_workbook(excel_file)
    sheet = workbook.active


    #for suite in json_info:
    json_suites=json_info#["Test_Suites"][0]
    #df.loc[df['Case identifier'] == json_suites["Suite_id"], 'Result'] = json_suites["result"]
    for tsu in json_suites["Test_Suites"]:
        df.loc[df['Case identifier'] == tsu["Suite_id"], 'Result'] = tsu["result"]
        for tc in tsu["Test_Cases"]:
            df.loc[df['Case identifier'] == tc["Case_id"], 'Result'] = tc["result"]
            for ts in tc["Test_Steps"]:
                i=df.index[df['Case identifier'] == ts["Step_id"]].tolist()[0]
                for ac in ts["Test_Actions"]:
                    df.loc[i, 'Result'] = ac
                    i=i+1

    for i, valor in enumerate(df["Result"], start=1):
        sheet.cell(row=i + 1, column=6, value=valor)
        
    workbook.save("Protocol_Dxxxxxx_RUN_"+str(json_info["Run"])+".xlsx")

if __name__ == "__main__":
    write_results("DO_prueba.json","D0000006245_Propulsion Management Functional Factory Type Test_Ed.A.xlsx")







































#if file_path:
#    # Load the selected Excel file
#    df = pd.read_excel(file_path, converters={"Case identifier": devolver_text,"Action Description": devolver_text})
# 
#    # Unmerge cells in column C, keeping values in the first row
#    df['Result Description'] = df['Result Description'].ffill()
#    df['Case identifier'] = df['Case identifier'].astype(str)
#    #for row in df['Result Description']:
#    #print(df.values)
#
#    df["Empty_Vars"] = df["Variables"].isna()
#    df["Empty_Res"] = df["Result Description"].isna()
#    df["Empty_Act"] = df["Action Description"].isna()
#
#    df['Type'] = np.where(df['Empty_Vars'] & df['Empty_Act'], 'ACA', 'AMA')
#
#    print(df)
#
#    test_suite={"id": "5",
#                "Test_Cases":[]}
#    test_cases_array=[]
#
#    arr=df.to_numpy()
#
#    #5 empty vars
#    #6 Empty Res
#    #7 Empty Act
#
#    flag_case_descript=0
#    flag_case_initials=0
#    is_step=0
#    is_case=0
#    test_step={}
#    test_steps_list=[]
#    test_action_list=[]
#    test_cases_list=[]
#
#    for line in arr:
#        if not("." in line.item(1)) and (not line.item(1)=="nan"):
#            is_suite=1
#            test_case={
#                "name":line.item(2),
#                "id": line.item(1),
#                "description":"",
#                "initial_conditions":[],
#                "test_steps":[]
#            }
#            flag_case_descript=1
#
#        elif flag_case_descript==1:
#            test_case.update({"description":line.item(2)})
#            flag_case_descript=0
#            flag_case_initials=1
#
#        elif flag_case_initials==1:
#            test_case.update({"initial_conditions":line.item(2).splitlines()[1:]})
#            flag_case_initials=0
#        
#        elif not (line.item(7)):
#                is_step=1
#                test_action_list=[]
#                test_step={"id":line.item(1),
#                           "test_actions":[]}
#                test_action_list.append({"text": line.item(2),"type":"MFA"})
#                test_step["test_actions"]=test_action_list
#            
#
#        if not (line.item(5)):
#            test_action_list.append({"text": line.item(3),"type":"ACA","variables":line.item(4).splitlines()})
#            test_step["test_actions"]=test_action_list
#            
#        elif not (line.item(6)):   
#            test_action_list.append({"text": line.item(3),"type":"MCA"})
#            test_step["test_actions"]=test_action_list
#        
#        
#        if is_step==1:
#            is_step=0
#            test_steps_list.append(test_step)
#            test_case["test_steps"]=test_steps_list
#
#        
#        #if is_case==1:
#        #    is_suite=0
#        #    test_cases_list.append(test_case)
#        #    test_suite["Test_Cases"]= test_cases_list
#    
#    test_cases_list.append(test_case)
#    test_suite["Test_Cases"]= test_cases_list
#
#    with open("C://Users//17940//Python_Testing//CAF_Sequencer//sequencer//structure.json", "w") as outfile:
#        json.dump(test_suite, outfile, indent = 4)
#
#
#    #json_str = json.dumps(test_step, indent=4)
#    #print(json_str)  
#
#
#        #elif not line.item(2) and line.item(2)
#
#
#           
#
#  
#    
#    #arr=df.to_numpy()
#    #x=arr.item(2,2)
#    #print(x)
#    
#    
#    
#    
#    
#    
#    
#    
#    
#    
#    
#    
#    
#    
#    #df.to_excel('modified_excel_file.xlsx', index=False)
# 
#    #print("File processed successfully!")
#else:
#    print("No file selected.")


